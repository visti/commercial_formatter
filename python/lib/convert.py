#!/usr/bin/env python3
import sys
import os
import csv
import shutil
from pathlib import Path
from openpyxl import load_workbook

# Simple ANSI colors
USE_COLORS = hasattr(sys.stdout, "isatty") and sys.stdout.isatty() and not os.environ.get("NO_COLOR")
DIM = "\033[2m" if USE_COLORS else ""
CYAN = "\033[96m" if USE_COLORS else ""
GREEN = "\033[92m" if USE_COLORS else ""
RED = "\033[91m" if USE_COLORS else ""
RESET = "\033[0m" if USE_COLORS else ""


def convert_xlsx_to_csv(xlsx_path):
    """Convert an XLSX file to CSV format."""
    # Load workbook with data_only=True to get evaluated cell values
    wb = load_workbook(xlsx_path, data_only=True)
    sheet = wb.active

    csv_path = os.path.splitext(xlsx_path)[0] + ".csv"

    with open(csv_path, "w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file, delimiter=";")
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)

    xlsx_name = Path(xlsx_path).name
    csv_name = Path(csv_path).name
    print(f"  {GREEN}✓{RESET} {xlsx_name} → {CYAN}{csv_name}{RESET}")

def main():
    if len(sys.argv) < 2:
        print("Usage: python convert.py file1.xlsx file2.xlsx ...")
        sys.exit(1)

    files = sys.argv[1:]
    print(f"{DIM}Converting {len(files)} XLSX file(s)...{RESET}")

    for file_path in files:
        if os.path.isfile(file_path):
            try:
                convert_xlsx_to_csv(file_path)

                # Move original to backup folder
                file_dir = os.path.dirname(file_path) or "."
                bak_folder = os.path.join(file_dir, "bak")
                os.makedirs(bak_folder, exist_ok=True)

                dest_path = os.path.join(bak_folder, os.path.basename(file_path))
                shutil.move(file_path, dest_path)
                print(f"    {DIM}(moved original to bak/){RESET}")
            except Exception as e:
                print(f"  {RED}✗{RESET} Error: {e}")
        else:
            print(f"  {RED}✗{RESET} File not found: {file_path}")

if __name__ == "__main__":
    main()

