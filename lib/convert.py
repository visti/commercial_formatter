#!/usr/bin/env python3
import sys
import os
import csv
import shutil
from openpyxl import load_workbook

def convert_xlsx_to_csv(xlsx_path):
    # Load workbook with data_only=True to get evaluated cell values
    wb = load_workbook(xlsx_path, data_only=True)
    # Use the active (first) sheet
    sheet = wb.active
    
    # Construct the CSV file name by replacing the .xlsx extension with .csv
    csv_path = os.path.splitext(xlsx_path)[0] + ".csv"
    
    # Open CSV file for writing
    with open(csv_path, "w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file, delimiter=";")
        # Iterate over rows in the active sheet and write them to CSV
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)
    
    print(f"Converted '{xlsx_path}' to '{csv_path}'")

def main():
    if len(sys.argv) < 2:
        print("Usage: python convert.py file1.xlsx file2.xlsx ...")
        sys.exit(1)
    
    # Process each XLSX file provided in the command-line arguments
    for file_path in sys.argv[1:]:
        if os.path.isfile(file_path):
            try:
                convert_xlsx_to_csv(file_path)
                
                # Determine the directory of the original file
                file_dir = os.path.dirname(file_path)
                # Create the backup folder "bak" within the same directory
                bak_folder = os.path.join(file_dir, "bak")
                os.makedirs(bak_folder, exist_ok=True)
                
                # Build the destination path inside the backup folder
                dest_path = os.path.join(bak_folder, os.path.basename(file_path))
                # Move the original XLSX file to the backup folder
                shutil.move(file_path, dest_path)
                print(f"Moved '{file_path}' to '{dest_path}'")
            except Exception as e:
                print(f"Error converting '{file_path}': {e}")
        else:
            print(f"File not found: '{file_path}'")

if __name__ == "__main__":
    main()

